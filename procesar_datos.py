import openpyxl
import json
import math
import os

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "Analisis Piloto Ev Patrocinadores.xlsx")
DATA_DIR = os.path.join(os.path.dirname(__file__), "data")

def fix_encoding(s):
    if not isinstance(s, str):
        return s
    try:
        return s.encode('latin-1').decode('utf-8')
    except Exception:
        return s

def mean(vals):
    v = [x for x in vals if x is not None]
    return round(sum(v) / len(v), 2) if v else None

def std(vals):
    v = [x for x in vals if x is not None]
    if len(v) < 2:
        return None
    m = sum(v) / len(v)
    variance = sum((x - m) ** 2 for x in v) / len(v)
    return round(math.sqrt(variance), 2)

def stats(vals):
    v = [x for x in vals if x is not None and isinstance(x, (int, float))]
    if not v:
        return {"promedio": None, "std": None, "min": None, "max": None, "n": 0, "alerta_n": True}
    return {
        "promedio": round(sum(v) / len(v), 2),
        "std": std(v),
        "min": min(v),
        "max": max(v),
        "n": len(v),
        "alerta_n": len(v) < 2
    }

def load_correo_region(wb):
    ws = wb["CorreoRegion"]
    rows = list(ws.iter_rows(values_only=True))[1:]
    mapping = {}
    for r in rows:
        if r[0] and r[1]:
            email = fix_encoding(str(r[0])).strip().lower()
            region = fix_encoding(str(r[1])).strip()
            mapping[email] = region
    return mapping

def normalize_entity(name):
    if not name:
        return name
    return fix_encoding(str(name)).strip()

def get_rol(email, correo_region):
    region = correo_region.get(email.lower(), "")
    return "Nacional" if region == "Nacional" else "Regional"

def load_formalizacion(wb, correo_region):
    ws = wb["Formalizacion"]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    q_labels = [
        "El convenio se firmó dentro del plazo establecido y conforme a las instrucciones enviadas por Corfo.",
        "Las garantías fueron entregadas en los plazos establecidos y conforme a las instrucciones enviadas por Corfo.",
        "La entidad se encuentra al día con los requisitos necesarios para formalizar y realizar el pago correspondiente.",
        "La comunicación con la entidad es fluida y permitió resolver dudas y requerimientos durante el proceso de formalización.",
        "En general, la entidad cumplió adecuadamente con el proceso completo de formalización de este proyecto."
    ]
    records = []
    for r in rows[1:]:
        email = fix_encoding(str(r[0])).strip() if r[0] else ""
        entity = normalize_entity(r[1])
        scores = [r[2], r[3], r[4], r[5], r[6]]
        # La región se toma del mapeo autoritativo CorreoRegion (completo);
        # la columna de la hoja queda solo como respaldo (suele venir vacía).
        region_sheet = fix_encoding(str(r[7])).strip() if r[7] else ""
        region = correo_region.get(email.lower(), "") or region_sheet
        rol = get_rol(email, correo_region)
        records.append({
            "email": email,
            "entidad": entity,
            "scores": scores,
            "region": region,
            "rol": rol
        })
    return records, q_labels

def load_seguimiento(wb, correo_region):
    ws = wb["Seguimiento"]
    rows = list(ws.iter_rows(values_only=True))
    q_labels = [
        "La entidad presentó los reportes de avance dentro de los plazos establecidos.",
        "La información contenida en los reportes fue clara y completa.",
        "El soporte técnico brindado por la entidad permitió resolver oportunamente las incidencias del proyecto.",
        "La entidad gestionó de manera proactiva los riesgos y desvíos identificados.",
        "La comunicación de la entidad con el equipo técnico de Corfo fue fluida y eficiente.",
        "En general, la calidad de la supervisión y monitoreo de la ejecución por parte de la entidad fue satisfactoria."
    ]
    records = []
    for r in rows[1:]:
        email = fix_encoding(str(r[0])).strip() if r[0] else ""
        entity = normalize_entity(r[1])
        scores = [r[2], r[3], r[4], r[5], r[6], r[7]]
        # La región se toma del mapeo autoritativo CorreoRegion (completo);
        # la columna de la hoja queda solo como respaldo (suele venir vacía).
        region_sheet = fix_encoding(str(r[8])).strip() if r[8] else ""
        region = correo_region.get(email.lower(), "") or region_sheet
        rol = get_rol(email, correo_region)
        records.append({
            "email": email,
            "entidad": entity,
            "scores": scores,
            "region": region,
            "rol": rol
        })
    return records, q_labels

def load_conteo(wb, correo_region):
    ws = wb["Conteo respuestas"]
    rows = list(ws.iter_rows(values_only=True))[1:]
    records = []
    for r in rows:
        if not r[0]:
            continue
        email = fix_encoding(str(r[0])).strip()
        entity = normalize_entity(r[2])
        rol_survey = fix_encoding(str(r[3])).strip() if r[3] else ""
        contestada = fix_encoding(str(r[5])).strip() if r[5] else "No"
        # Skip Sostenibilidad y Género
        if "Sostenibilidad" in rol_survey or "Género" in rol_survey or "Genero" in rol_survey:
            continue
        # Normalize survey type
        if "Ejecucion" in rol_survey or "Monitoreo" in rol_survey:
            tipo = "Seguimiento"
        elif "Formalizacion" in rol_survey or "Formalización" in rol_survey:
            tipo = "Formalizacion"
        else:
            continue
        rol = get_rol(email, correo_region)
        region = correo_region.get(email.lower(), "Sin región")
        records.append({
            "email": email,
            "entidad": entity,
            "tipo": tipo,
            "contestada": contestada == "Si",
            "rol": rol,
            "region": region
        })
    return records

def build_questions_json(form_records, form_labels, seg_records, seg_labels):
    def build_question_stats(records, q_idx, q_label, q_id):
        all_scores = [r["scores"][q_idx] for r in records if r["scores"][q_idx] is not None]
        by_rol = {}
        for rol in ["Nacional", "Regional"]:
            vs = [r["scores"][q_idx] for r in records if r["rol"] == rol and r["scores"][q_idx] is not None]
            by_rol[rol] = stats(vs)
        entities = sorted(set(r["entidad"] for r in records if r["entidad"]))
        by_entidad = {}
        for e in entities:
            vs = [r["scores"][q_idx] for r in records if r["entidad"] == e and r["scores"][q_idx] is not None]
            by_entidad[e] = stats(vs)
        regions = sorted(set(r["region"] for r in records if r["region"]))
        by_region = {}
        for reg in regions:
            vs = [r["scores"][q_idx] for r in records if r["region"] == reg and r["scores"][q_idx] is not None]
            by_region[reg] = stats(vs)
        s = stats(all_scores)
        s["id"] = q_id
        s["texto"] = q_label
        s["por_rol"] = by_rol
        s["por_entidad"] = by_entidad
        s["por_region"] = by_region
        return s

    form_qs = [build_question_stats(form_records, i, form_labels[i], f"F{i+1}") for i in range(len(form_labels))]
    seg_qs = [build_question_stats(seg_records, i, seg_labels[i], f"S{i+1}") for i in range(len(seg_labels))]
    return {"formalizacion": form_qs, "seguimiento": seg_qs}

def build_entities_json(form_records, seg_records):
    def entity_stats(records):
        entities = sorted(set(r["entidad"] for r in records if r["entidad"]))
        result = []
        for e in entities:
            scores_all = []
            for r in records:
                if r["entidad"] == e:
                    scores_all.extend([s for s in r["scores"] if s is not None])
            s = stats(scores_all)
            s["entidad"] = e
            # per question averages
            n_q = len(records[0]["scores"]) if records else 0
            per_q = {}
            for qi in range(n_q):
                vs = [r["scores"][qi] for r in records if r["entidad"] == e and r["scores"][qi] is not None]
                per_q[f"Q{qi+1}"] = round(sum(vs)/len(vs), 2) if vs else None
            s["por_pregunta"] = per_q
            result.append(s)
        result.sort(key=lambda x: x["promedio"] if x["promedio"] else 0, reverse=True)
        return result

    return {
        "formalizacion": entity_stats(form_records),
        "seguimiento": entity_stats(seg_records)
    }

def build_regions_json(form_records, seg_records):
    all_regions = sorted(set(
        list(set(r["region"] for r in form_records if r["region"])) +
        list(set(r["region"] for r in seg_records if r["region"]))
    ))
    result = []
    for reg in all_regions:
        fr = [r for r in form_records if r["region"] == reg]
        sr = [r for r in seg_records if r["region"] == reg]
        f_scores = [s for r in fr for s in r["scores"] if s is not None]
        s_scores = [s for r in sr for s in r["scores"] if s is not None]
        f_prom = round(sum(f_scores)/len(f_scores), 2) if f_scores else None
        s_prom = round(sum(s_scores)/len(s_scores), 2) if s_scores else None
        gen = None
        if f_prom is not None and s_prom is not None:
            gen = round((f_prom + s_prom) / 2, 2)
        elif f_prom is not None:
            gen = f_prom
        elif s_prom is not None:
            gen = s_prom
        result.append({
            "region": reg,
            "promedio_formalizacion": f_prom,
            "promedio_seguimiento": s_prom,
            "promedio_general": gen,
            "n_formalizacion": len(fr),
            "n_seguimiento": len(sr),
            "alerta_n": len(fr) < 2 or len(sr) < 2
        })
    return result

def build_heatmap_json(form_records, seg_records, form_labels, seg_labels):
    def build_heatmap(records, labels, prefix):
        entities = sorted(set(r["entidad"] for r in records if r["entidad"]))
        n_q = len(labels)
        data = []
        for e in entities:
            row = []
            for qi in range(n_q):
                vs = [r["scores"][qi] for r in records if r["entidad"] == e and r["scores"][qi] is not None]
                row.append(round(sum(vs)/len(vs), 2) if vs else None)
            data.append(row)
        return {
            "entidades": entities,
            "preguntas": [f"{prefix}{i+1}" for i in range(n_q)],
            "etiquetas": labels,
            "data": data
        }
    return {
        "formalizacion": build_heatmap(form_records, form_labels, "F"),
        "seguimiento": build_heatmap(seg_records, seg_labels, "S")
    }

def build_radar_json(form_records, seg_records, form_labels, seg_labels):
    def build_radar(records, labels, prefix):
        result = {}
        for rol in ["Nacional", "Regional"]:
            r_recs = [r for r in records if r["rol"] == rol]
            rol_data = {}
            for qi, label in enumerate(labels):
                vs = [r["scores"][qi] for r in r_recs if r["scores"][qi] is not None]
                rol_data[f"{prefix}{qi+1}"] = round(sum(vs)/len(vs), 2) if vs else None
            result[rol] = rol_data
        return result
    return {
        "formalizacion": build_radar(form_records, form_labels, "F"),
        "seguimiento": build_radar(seg_records, seg_labels, "S"),
        "etiquetas": {
            "formalizacion": {f"F{i+1}": l for i, l in enumerate(form_labels)},
            "seguimiento": {f"S{i+1}": l for i, l in enumerate(seg_labels)}
        }
    }

def build_response_rate_json(conteo_records, correo_region):
    def rate_for(records):
        total = len(records)
        contestadas = sum(1 for r in records if r["contestada"])
        return {
            "total": total,
            "contestadas": contestadas,
            "tasa": round(contestadas / total, 4) if total else 0
        }

    # por tipo
    form_c = [r for r in conteo_records if r["tipo"] == "Formalizacion"]
    seg_c = [r for r in conteo_records if r["tipo"] == "Seguimiento"]

    # por rol
    por_rol = {}
    for rol in ["Nacional", "Regional"]:
        recs = [r for r in conteo_records if r["rol"] == rol]
        por_rol[rol] = rate_for(recs)

    # por entidad
    entities = sorted(set(r["entidad"] for r in conteo_records))
    por_entidad = {}
    for e in entities:
        recs = [r for r in conteo_records if r["entidad"] == e]
        por_entidad[e] = {
            "general": rate_for(recs),
            "formalizacion": rate_for([r for r in recs if r["tipo"] == "Formalizacion"]),
            "seguimiento": rate_for([r for r in recs if r["tipo"] == "Seguimiento"])
        }

    # por region
    regions = sorted(set(r["region"] for r in conteo_records if r["region"]))
    por_region = {}
    for reg in regions:
        recs = [r for r in conteo_records if r["region"] == reg]
        por_region[reg] = {
            "general": rate_for(recs),
            "formalizacion": rate_for([r for r in recs if r["tipo"] == "Formalizacion"]),
            "seguimiento": rate_for([r for r in recs if r["tipo"] == "Seguimiento"])
        }

    return {
        "total": rate_for(conteo_records),
        "formalizacion": rate_for(form_c),
        "seguimiento": rate_for(seg_c),
        "por_rol": por_rol,
        "por_entidad": por_entidad,
        "por_region": por_region
    }

def build_summary_json(form_records, seg_records, form_labels, seg_labels, response_rate):
    all_f = [s for r in form_records for s in r["scores"] if s is not None]
    all_s = [s for r in seg_records for s in r["scores"] if s is not None]
    prom_f = round(sum(all_f)/len(all_f), 2) if all_f else None
    prom_s = round(sum(all_s)/len(all_s), 2) if all_s else None
    prom_gen = round((prom_f + prom_s) / 2, 2) if prom_f and prom_s else None

    # Find min/max question
    umbral = 5.0
    alertas = []

    # Questions below threshold
    for qi, label in enumerate(form_labels):
        vs = [r["scores"][qi] for r in form_records if r["scores"][qi] is not None]
        if vs:
            avg = sum(vs)/len(vs)
            if avg < umbral:
                alertas.append({
                    "tipo": "bajo_umbral",
                    "ambito": "Formalización",
                    "pregunta": f"F{qi+1}",
                    "texto": label[:80] + "...",
                    "promedio": round(avg, 2)
                })
    for qi, label in enumerate(seg_labels):
        vs = [r["scores"][qi] for r in seg_records if r["scores"][qi] is not None]
        if vs:
            avg = sum(vs)/len(vs)
            if avg < umbral:
                alertas.append({
                    "tipo": "bajo_umbral",
                    "ambito": "Seguimiento",
                    "pregunta": f"S{qi+1}",
                    "texto": label[:80] + "...",
                    "promedio": round(avg, 2)
                })

    # Low response count entities
    entities = sorted(set(r["entidad"] for r in form_records + seg_records if r["entidad"]))
    for e in entities:
        nf = len([r for r in form_records if r["entidad"] == e])
        ns = len([r for r in seg_records if r["entidad"] == e])
        if nf < 2:
            alertas.append({"tipo": "pocas_respuestas", "entidad": e, "ambito": "Formalización", "n": nf})
        if ns < 2:
            alertas.append({"tipo": "pocas_respuestas", "entidad": e, "ambito": "Seguimiento", "n": ns})

    return {
        "promedio_general": prom_gen,
        "promedio_formalizacion": prom_f,
        "promedio_seguimiento": prom_s,
        "tasa_respuesta_general": response_rate["total"]["tasa"],
        "total_evaluaciones": response_rate["total"]["total"],
        "contestadas": response_rate["total"]["contestadas"],
        "umbral": umbral,
        "alertas": alertas,
        "n_formalizacion": len(form_records),
        "n_seguimiento": len(seg_records),
        "entidades": sorted(set(r["entidad"] for r in form_records + seg_records if r.get("entidad"))),
        "regiones": sorted(set(r["region"] for r in form_records + seg_records if r.get("region")))
    }

def main():
    import shutil
    # Copia temporal SIEMPRE refrescada desde el Excel actual (evita el bloqueo de
    # OneDrive/Excel sin cachear datos viejos).
    tmp = os.path.join(os.environ.get("TEMP", "/tmp"), "AnalisisPiloto_proc.xlsx")
    shutil.copy2(EXCEL_PATH, tmp)

    wb = openpyxl.load_workbook(tmp, read_only=True, data_only=True)
    correo_region = load_correo_region(wb)

    form_records, form_labels = load_formalizacion(wb, correo_region)
    seg_records, seg_labels = load_seguimiento(wb, correo_region)
    conteo_records = load_conteo(wb, correo_region)

    os.makedirs(DATA_DIR, exist_ok=True)

    questions = build_questions_json(form_records, form_labels, seg_records, seg_labels)
    entities = build_entities_json(form_records, seg_records)
    regions = build_regions_json(form_records, seg_records)
    heatmap = build_heatmap_json(form_records, seg_records, form_labels, seg_labels)
    radar = build_radar_json(form_records, seg_records, form_labels, seg_labels)
    response_rate = build_response_rate_json(conteo_records, correo_region)
    summary = build_summary_json(form_records, seg_records, form_labels, seg_labels, response_rate)

    # Anonymized records for client-side filtering
    records_data = {
        "formalizacion": [
            {"entidad": r["entidad"], "region": r["region"], "rol": r["rol"], "scores": r["scores"]}
            for r in form_records if r.get("entidad")
        ],
        "seguimiento": [
            {"entidad": r["entidad"], "region": r["region"], "rol": r["rol"], "scores": r["scores"]}
            for r in seg_records if r.get("entidad")
        ],
        "preguntas_formalizacion": form_labels,
        "preguntas_seguimiento": seg_labels
    }

    files = {
        "summary.json": summary,
        "questions.json": questions,
        "entities.json": entities,
        "regions.json": regions,
        "heatmap.json": heatmap,
        "radar.json": radar,
        "response_rate.json": response_rate,
        "records.json": records_data
    }

    for fname, data in files.items():
        path = os.path.join(DATA_DIR, fname)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f"Generado: {path}")

    # Generate data.js for GitHub Pages / server mode
    js_path = os.path.join(DATA_DIR, "data.js")
    inline_js = (
        "// Auto-generado por procesar_datos.py — no editar manualmente\n"
        f"window.DATA_RECORDS = {json.dumps(records_data, ensure_ascii=False)};\n"
        f"window.DATA_SUMMARY = {json.dumps(summary, ensure_ascii=False)};\n"
        f"window.DATA_RESPONSE_RATE = {json.dumps(response_rate, ensure_ascii=False)};\n"
    )
    with open(js_path, "w", encoding="utf-8") as f:
        f.write(inline_js)
    print(f"Generado: {js_path}")

    # Inject data directly into index.html for offline use (no server required)
    html_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "index.html")
    BEGIN = "<!-- DATA_BEGIN -->"
    END = "<!-- DATA_END -->"
    if os.path.exists(html_path):
        with open(html_path, "r", encoding="utf-8") as f:
            html = f.read()
        if BEGIN in html and END in html:
            before = html.split(BEGIN)[0]
            after = html.split(END)[1]
            html_new = (
                before
                + BEGIN + "\n"
                + "<script>\n" + inline_js + "</script>\n"
                + END
                + after
            )
            with open(html_path, "w", encoding="utf-8") as f:
                f.write(html_new)
            print(f"Datos inyectados en: {html_path}")
        else:
            print("AVISO: No se encontraron marcadores DATA_BEGIN/DATA_END en index.html")
    else:
        print(f"AVISO: No se encontró index.html en {html_path}")

    print("\nResumen:")
    print(f"  Promedio general: {summary['promedio_general']}")
    print(f"  Promedio Formalización: {summary['promedio_formalizacion']}")
    print(f"  Promedio Seguimiento: {summary['promedio_seguimiento']}")
    print(f"  Tasa de respuesta: {summary['tasa_respuesta_general']:.1%}")
    print(f"  Total evaluaciones: {summary['total_evaluaciones']}")
    print(f"  Alertas: {len(summary['alertas'])}")
    print(f"  Entidades: {summary['entidades']}")

if __name__ == "__main__":
    main()
